import os
from django.contrib import auth
from django.shortcuts import render
from h11 import Data
from account.models import User
from crawler.models import *
from config.models import PlatformTargetItem, CollectTargetItem
from config.serializers import PlatformTargetItemSerializer, CollectTargetItemSerializer
from dataprocess.functions import export_datareport, import_datareport, import_total
from django.views.decorators.csrf import csrf_exempt

from .serializers import *
from .models import *
from django.http.response import JsonResponse
from rest_framework.parsers import JSONParser 
from rest_framework import status
from django.views.decorators.csrf import csrf_exempt
from utils.decorators import login_required
from utils.api import APIView, validate_serializer

from django.shortcuts import render
from django.http import HttpResponse

from datetime import datetime
import openpyxl
from openpyxl.writer.excel import save_virtual_workbook
from .resources import *
from .models import *
import logging

formatter = logging.Formatter('[%(asctime)s] - [%(levelname)s] - [%(name)s:%(lineno)d]  - %(message)s', '%Y-%m-%d %H:%M:%S')
serverlogger = logging.getLogger(__name__)
userlogger = logging.getLogger("HTTP-Method")

trfh = logging.handlers.TimedRotatingFileHandler(
    filename = os.path.join("../backend/data/log/user", f"{datetime.today().strftime('%Y-%m-%d')}.log"),
    when = "midnight",
    interval=1,
    encoding="utf-8",
)
trfh.setFormatter(formatter)
trfh.setLevel(logging.INFO)
userlogger.addHandler(trfh)
userlogger.setLevel(logging.DEBUG)

DataModels = {
        "youtube": SocialbladeYoutube,
        "tiktok": SocialbladeTiktok,
        "twitter": SocialbladeTwitter,
        "twitter2": SocialbladeTwitter2,
        "weverse": Weverse,
        "instagram": CrowdtangleInstagram,
        "facebook": CrowdtangleFacebook,
        "vlive": Vlive,
        "melon": Melon,
        "spotify": Spotify,
}

# login check using cookie
def logincheck(request):
    # 로그인 정보를 받기 위해 cookie사용
    username = request.COOKIES.get('username')
    if username is not None:
        if User.objects.filter(username=username).exists():
            # 이미 존재하는 username일때만 로그인
            user = User.objects.filter(username=username).first()
            auth.login(request, user)
    return request

# Create your views here.
def base(request):
    values = {
      'first_depth' : '데이터 리포트',
    }
    request = logincheck(request)
    return render(request, 'dataprocess/main.html',values)

@csrf_exempt
def daily(request):
    if request.method == 'GET':
        '''
        general page
        '''
        platforms = Platform.objects.all() #get all platform info from db
        values = {
            'first_depth' : '데이터 리포트',
            'second_depth': '일별 리포트',
            'platforms': platforms
        }
        request = logincheck(request)
        return render(request, 'dataprocess/daily.html',values)
    else:
        type = request.POST['type']
        if type == 'import':
            '''
            import from excel
            '''
            import_file = request.FILES['importData']
            # platform = request.GET.get('platform', None)
            excel_import_date = request.POST.get('excel_import_date', None) # 0000-0-0 형태

            wb = openpyxl.load_workbook(import_file)
            sheets = wb.sheetnames
            worksheet = wb[sheets[0]]
            import_datareport(worksheet, excel_import_date)
            platforms = Platform.objects.all() #get all platform info from db
            values = {
                'first_depth' : '데이터 리포트',
                'second_depth': '일별 리포트',
                'platforms': platforms,
                'alert': 'Successfully save to DB!'
                }
            request = logincheck(request)
            return render(request, 'dataprocess/daily.html',values)
        elif type == 'export':
            '''
            export to excel
            '''
            excel_export_type = request.POST.get('excel_export_days', None) # 누적 or 기간별
            excel_export_start_date = request.POST.get('excel_export_start_date', None) # 0000-0-0 형태
            excel_export_end_date = request.POST.get('excel_export_end_date', None) # 0000-0-0 형태
            book = export_datareport(excel_export_type, excel_export_start_date, excel_export_end_date)
            today_date = datetime.datetime.today()
            filename = 'datareport%s %s-%s-%s.xlsx'%excel_export_type, (today_date.year, today_date.month, today_date.day)
            response = HttpResponse(content=save_virtual_workbook(book), content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename='+filename
            return response
        elif type == 'import2':
            '''
            import2 from excel
            '''
            import_file = request.FILES['importData2']
            wb = openpyxl.load_workbook(import_file)
            sheets = wb.sheetnames
            worksheet = wb[sheets[0]]
            import_total(worksheet)
            platforms = Platform.objects.all() #get all platform info from db
            values = {
                'first_depth' : '데이터 리포트',
                'second_depth': '일별 리포트',
                'platforms': platforms,
                'alert': 'Successfully save to DB!'
                }
            request = logincheck(request)
            return render(request, 'dataprocess/daily.html',values)
    
def platform(request):
     '''
        general page
        '''
     values = {
        'first_depth' : '플랫폼 관리',
        'second_depth': '플랫폼 관리'
    }
     request = logincheck(request)
     return render(request, 'dataprocess/platform.html',values)


# def excel(request):
#     if request.method == 'POST':
#         type = request.POST['type']
#         if type == 'import':
#             '''
#             import from excel
#             '''
#             import_file = request.FILES['importData']
#             wb = openpyxl.load_workbook(import_file)
#             sheets = wb.sheetnames
#             worksheet = wb[sheets[0]]
#             import_datareport(worksheet)
#             platforms = Platform.objects.all() #get all platform info from db
#             values = {
#                 'first_depth' : '데이터 리포트',
#                 'second_depth': '일별 리포트',
#                 'platforms': platforms
#                 }
#             request = logincheck(request)
#             return render(request, 'dataprocess/daily.html',values)
#         elif type == 'export':
#             '''
#             export to excel
#             '''
#             book = export_datareport()
#             today_date = datetime.datetime.today()
#             filename = 'datareport%s-%s-%s.xlsx'%(today_date.year, today_date.month, today_date.day)
#             response = HttpResponse(content=save_virtual_workbook(book), content_type='application/vnd.ms-excel')
#             response['Content-Disposition'] = 'attachment; filename='+filename
#             return response
#         elif type == 'import2':
#             '''
#             import tmp from excel
#             '''
#             import_file = request.FILES['importData2']
#             wb = openpyxl.load_workbook(import_file)
#             sheets = wb.sheetnames
#             worksheet = wb[sheets[0]]
#             import_total(worksheet)
#             platforms = Platform.objects.all() #get all platform info from db
#             values = {
#                 'first_depth' : '데이터 리포트',
#                 'second_depth': '일별 리포트',
#                 'platforms': platforms
#                 }
#             request = logincheck(request)
#             return render(request, 'dataprocess/daily.html',values)

def artist(request):
    artists = Artist.objects.all()
    values = {
      'first_depth' : '아티스트 관리',
      'second_depth': '데이터 URL 관리',
      'artists': artists
    }
    request = logincheck(request)
    return render(request, 'dataprocess/artist.html',values)

@csrf_exempt
def artist_add(request):
    platforms = Platform.objects.all()
    values = {
      'first_depth' : '아티스트 관리',
      'second_depth': '데이터 URL 관리',
      'platforms' : platforms
    }
    request = logincheck(request)
    return render(request, 'dataprocess/artist_add.html',values)

def monitering(request):
    platforms = Platform.objects.all()
    values = {
      'first_depth' : '모니터링 관리',
      'second_depth': '모니터링',
      'platforms' : platforms
    }
    return render(request, 'dataprocess/monitering.html',values)


def login(request):
    values = {
      'first_depth' : '로그인'
    }
    request = logincheck(request)
    return render(request, 'dataprocess/login.html',values)

def platform_info(request):
    if request.method == 'GET':
        platform = request.GET.get('platform', None)
        try:
            platform_objects = Platform.objects.get(name = platform)
            
            if platform_objects.exists():
                platform_objects_values = platform_objects.values()
                platform_id = platform_objects_values['id']
                collecttargets = CollectTarget.objects.filter(platform = platform_id)
                collecttargets = collecttargets.values()
                platform_set = set()
                platform_list = []
                for collecttarget in collecttargets:
                    platform_objects = CollectTargetItem.objects.filter(collect_target_id = collecttarget['id'])
                    platform_objects_values = platform_objects.values()
                    for p in platform_objects_values:
                        if p["target_name"] in platform_set:
                            continue
                        platform_set.add(p["target_name"])
                        platform_list.append(p)
                return JsonResponse(data={'success': True, 'data': platform_list})
            else:
                return JsonResponse(data={'success': True, 'data': []})
        except:
            return JsonResponse(status=400, data={'success': False})


class PlatformAPI(APIView):
    # @login_required
    def get(self, request):
        """
        Platform read api
        """
        try:
            platform_objects = Platform.objects.all()
            if platform_objects.exists():
                platform_objects_values = platform_objects.values()
                platform_datas = []
                for platform_value in platform_objects_values:
                    platform_datas.append(platform_value)
                return JsonResponse(data={'success': True, 'data': platform_datas})
            else:
                return JsonResponse(data={'success': True, 'data': []})
        except:
            return JsonResponse(status=400, data={'success': False})

    # @login_required
    def post(self, request):
        """
        Platform create api
        """
        try:
            platform_object = JSONParser().parse(request)
            platform_serializer = PlatformSerializer(data=platform_object)
            if platform_serializer.is_valid():
                # 1. platform 생성
                platform_serializer.save()

                #특정 아티스트 전용 collect_target 생성 시 사용할 코드
                # 2. 현재 존재하는 모든 artist에 대해 collect_target 생성 -> platform과 연결
                artist_objects = Artist.objects.all()
                artist_objects_values = artist_objects.values()
                for artist_objects_value in artist_objects_values:
                    collecttarget = CollectTarget(
                        platform_id = platform_serializer.data['id'],
                        artist_id = artist_objects_value['id']
                        )
                    collecttarget.save()
                    #3. 만든 collect target에 대해 수집항목들 생성
                    collecttarget_object = CollectTarget.objects.filter(platform = platform_serializer.data['id'],
                            artist = artist_objects_value['id'])
                    collecttarget_object = collecttarget_object.values()[0]
                    for collect_item in platform_object['collect_items']:
                        collect_item = CollectTargetItem(
                            collect_target_id = collecttarget_object['id'],
                            target_name = collect_item["target_name"],
                            xpath = collect_item["xpath"]
                        )
                        collect_item.save()
                
                #3. 플랫폼에 대한 platform target 생성
                # for collect_item in platform_object['collect_items']:
                #     collect_item = PlatformTargetItem(
                #         platform_id = platform_serializer.data['id'],
                #         target_name = collect_item["target_name"],
                #         xpath = collect_item["xpath"]
                #     )
                #     collect_item.save()

                return JsonResponse(data={'success': True, 'data': platform_serializer.data}, status=status.HTTP_201_CREATED)
            return JsonResponse(data={'success': False,'data': platform_serializer.errors}, status=400)
        except:
            return JsonResponse(data={'success': False}, status=400)

    # @login_required
    def put(self, request):
        """
        Platform update api
        """
        try:
            platform_list = JSONParser().parse(request)
            for platform_object in platform_list:
                platform_data = Platform.objects.get(id=platform_object["id"])
                if platform_data is None:
                    # 원래 없는 건 새로 저장
                    platform_serializer = PlatformSerializer(data=platform_object)
                    if platform_serializer.is_valid():
                        platform_serializer.save()
                else:
                    data = PlatformSerializer(platform_data).data
                    print(data)
                    past_name = data["name"]
                    past_url = data["url"]
                    cur_name = platform_object["name"]
                    cur_url = platform_object["url"]
                    platform_serializer = PlatformSerializer(platform_data, data=platform_object)
                    if platform_serializer.is_valid():
                        if past_name != cur_name:
                            userlogger.info(f"[CHANGE]: {past_name} -> {cur_name}")
                        if past_url != cur_url:
                            userlogger.info(f"[CHANGE]: {past_url} -> {cur_url}")
                        platform_serializer.save()
            return JsonResponse(data={'success': True}, status=status.HTTP_201_CREATED)
        except:
            return JsonResponse(data={'success': False}, status=400)

class ArtistAPI(APIView):
    # @login_required
    def get(self, request):
        """
        Artist read api
        """
        try:
            artist_objects = Artist.objects.all()
            if artist_objects.exists():
                artist_objects_values = artist_objects.values()
                artist_datas = []
                for artist_value in artist_objects_values:
                    artist_datas.append(artist_value)
                return JsonResponse(status=200,data={'success': True, 'data': artist_datas})
            else:
                return JsonResponse(data={'success': True, 'data': []})
        except:
            return JsonResponse(status=400, data={'success': False})

    # @login_required
    def post(self, request):
        """
        Artist create api
        """
        try:
            artist_object = JSONParser().parse(request)
            artist_serializer = ArtistSerializer(data=artist_object)
            if artist_serializer.is_valid():
                # 1. artist 생성
                artist_serializer.save()
                # 2. 현재 존재하는 모든 platform에 대해 collect_target 생성 -> artist와 연결
                platform_objects = Platform.objects.all()

                for obj in artist_object['urls']:
                    platform_id = Platform.objects.get(name = obj['platform_name']).id
                    artist_id = artist_serializer.data['id']
                    target_url = obj['url1']
                    target_url_2 = obj['url2']
                    collecttarget = CollectTarget(
                        platform_id = platform_id,
                        artist_id = artist_id,
                        target_url = target_url,
                        target_url_2 = target_url_2
                    )
                    collecttarget.save()
                
    
                return JsonResponse(data={'success': True, 'data': artist_serializer.data}, status=status.HTTP_201_CREATED)
            return JsonResponse(data={'success': False,'data': artist_serializer.errors}, status=400)
        except:
            return JsonResponse(data={'success': False}, status=400)

    # @login_required
    def put(self, request):
        """
        Artist update api
        """
        try:
            artist_list = JSONParser().parse(request)
            for artist_object in artist_list:
                artist_data = Artist.objects.get(id=artist_object["id"])
                data = ArtistSerializer(artist_data).data
                print(data)
                past_name = data["name"]
                past_num = data["member_num"]
                past_agency = data["agency"]
                cur_name = artist_object["name"]
                cur_num = artist_object["member_num"]
                cur_agnecy = artist_object["agnecy"]
                artist_serializer = ArtistSerializer(artist_data, data=artist_object)
                if artist_serializer.is_valid():
                    if past_name != cur_name:
                        userlogger.info(f"[CHANGE]: {past_name} -> {cur_name}")
                    if past_num != cur_num:
                        userlogger.info(f"[CHANGE]: {past_num} -> {cur_num}")
                    if past_agency != cur_agnecy:
                        userlogger.info(f"[CHANGE]: {past_agency} -> {cur_agnecy}")
                    artist_serializer.save()
                else:
                    return JsonResponse(data={'success': False,'data': artist_serializer.errors}, status=400)
            return JsonResponse(data={'success': True}, status=status.HTTP_201_CREATED)
        except:
            return JsonResponse(data={'success': False}, status=400)


class PlatformOfArtistAPI(APIView):
    # @login_required
    def get(self, request):
        """
        Platform of Artist read api
        """
        try:
            artist = request.GET.get('artist', None)
            # 해당 artist 찾기
            artist_object = Artist.objects.filter(name = artist)
            artist_object = artist_object.values()[0]
            # 해당 artist를 가지는 collect_target들 가져오기
            collecttarget_objects = CollectTarget.objects.filter(artist_id=artist_object['id'])
            if collecttarget_objects.exists():
                collecttarget_objects_values = collecttarget_objects.values()
                platform_datas = []
                for collecttarget_value in collecttarget_objects_values:
                    platform_object = Platform.objects.get(pk=collecttarget_value['platform_id'])
                    platform_datas.append({  
                        'artist_id':artist_object['id'],
                        'platform_id' : collecttarget_value['platform_id'],
                        'id': collecttarget_value['id'],
                        'name': platform_object.name,
                        'target_url':collecttarget_value['target_url'],
                        'target_url_2': collecttarget_value['target_url_2']
                    })
                return JsonResponse(data={'success': True, 'data': platform_datas})
            else:
                return JsonResponse(data={'success': True, 'data': []})
        except:
            return JsonResponse(status=400, data={'success': False})

    # @login_required
    def put(self, request):
        """
        Platform of Artist update api
        """
        try:
            collecttarget_list = JSONParser().parse(request)
            for collecttarget_object in collecttarget_list:
                data = CollectTarget.objects.get(id=collecttarget_object["id"])
                print(data)
                CollectTarget.objects.filter(pk=collecttarget_object['id']).update(target_url=collecttarget_object['target_url'])
                if collecttarget_object['target_url_2']:
                     CollectTarget.objects.filter(pk=collecttarget_object['id']).update(target_url_2=collecttarget_object['target_url_2'])
            return JsonResponse(data={'success': True}, status=status.HTTP_201_CREATED)
        except:
            return JsonResponse(data={'success': False}, status=400)


class CollectTargetItemAPI(APIView):
    # @login_required
    def get(self, request):
        """
        CollectTargetItem read api
        """
        try:
            artist = request.GET.get('artist', None)
            platform = request.GET.get('platform', None)
            # 해당 artist,platform 찾기
            artist_object = Artist.objects.filter(name = artist)
            artist_object = artist_object.values()[0]
            platform_object = Platform.objects.filter(name = platform)
            platform_object = platform_object.values()[0]
            # 해당 artist와 platform을 가지는 collect_target 가져오기
            collecttarget_objects = CollectTarget.objects.filter(artist_id=artist_object['id'], platform_id = platform_object['id'])
            if collecttarget_objects.exists():
                collecttargetitems_datas = []
                collecttarget_objects_value = collecttarget_objects.values()[0]
                collecttargetitmes_objects = CollectTargetItem.objects.filter(collect_target_id=collecttarget_objects_value['id'])
                collecttargetitmes_values = collecttargetitmes_objects.values()
                for collecttargetitmes_value in collecttargetitmes_values:
                    collecttargetitems_datas.append(collecttargetitmes_value)
                return JsonResponse(data={'success': True, 'data': collecttargetitems_datas})
            else:
                return JsonResponse(data={'success': True, 'data': []})
        except:
            return JsonResponse(status=400, data={'success': False})

    # @login_required
    def put(self, request):
        """
        CollectTargetItem update api
        """
        try:
            collecttargetitem = JSONParser().parse(request)
            collecttargetitem_list = collecttargetitem["items"]
            for collecttargetitem_object in collecttargetitem_list:
                # 여기 수정!!!!
                collecttargetitem_data = CollectTargetItem.objects.filter(id=collecttargetitem_object['id'],
                    target_name=collecttargetitem_object['target_name'],xpath=collecttargetitem_object['xpath']).first()
                # 없으면 새로 저장
                if collecttargetitem_data is None:
                    artist_object = Artist.objects.filter(name = collecttargetitem["artist"])
                    artist_object = artist_object.values()[0]
                    platform_object = Platform.objects.filter(name = collecttargetitem["platform"])
                    platform_object = platform_object.values()[0]
                    collecttarget_object = CollectTarget.objects.filter(artist_id=artist_object['id'], platform_id=platform_object['id'])
                    collecttarget_object = collecttarget_object.values()[0]
                    target_item_serializer = CollectTargetItemSerializer(data={
                        'collect_target': collecttarget_object['id'],
                        'target_name': collecttargetitem_object['target_name'],
                        'xpath': collecttargetitem_object['xpath']
                    })
                    if target_item_serializer.is_valid():
                        target_item_serializer.save()
                    else:
                        return JsonResponse(data={'success': False,'data': collecttargetitem_serializer.errors}, status=400)
                # 있으면 업데이트
                else:
                    collecttargetitem_serializer = CollectTargetItemSerializer(collecttargetitem_data, data=collecttargetitem_object)
                    if collecttargetitem_serializer.is_valid():
                        collecttargetitem_serializer.save()
# =======
#                     past_target_name = data["target_name"]
#                     past_xpath = data["xpath"]
#                     cur_target_name = collecttargetitem_object["target_name"]
#                     cur_xpath = collecttargetitem_object["xpath"]
#                     collecttargetitem_serializer = CollectTargetItemSerializer(data, data=collecttargetitem_object)
#                     if collecttargetitem_serializer.is_valid():
#                         collecttargetitem_serializer.save()
#                         if past_target_name != cur_target_name:
#                             userlogger.debug(f"{artist} - {platform} - {period}: {past_target_name} change to {cur_target_name}")
#                         if past_xpath != cur_xpath:
#                             userlogger.debug(f"{artist} - {platform} - {period}: {past_xpath} change to {cur_xpath}")
                    else:
                        return JsonResponse(data={'success': False,'data': collecttargetitem_serializer.errors}, status=400)
            return JsonResponse(data={'success': True}, status=status.HTTP_201_CREATED)
        except:
            return JsonResponse(data={'success': False}, status=400)

    def delete(self, request):
        """
        CollectTargetItem delete api
        """
        try:
            delete_id = JSONParser().parse(request)["id"]
            obj = CollectTargetItem.objects.filter(id = delete_id)
            obj.delete()
            return JsonResponse(data={'success': True}, status=status.HTTP_200_OK)
        except:
            return JsonResponse(data={'success': False}, status=400)

#platform collect target API 
class PlatformTargetItemAPI(APIView):
    # @login_required
    def get(self, request):
        """
        PlatformTargetItem read api
        """
        try:
            platform = request.GET.get('platform', None)
            # 해당 platform 찾기
            platform_object = Platform.objects.filter(name = platform)
            platform_object = platform_object.values()[0]
            # 해당 platform을 가지는 platform_target 가져오기
            collecttarget_objects = PlatformTargetItem.objects.filter(platform_id = platform_object['id'])
            if collecttarget_objects.exists():
                collecttargetitems_datas = []
                collecttarget_objects_value = collecttarget_objects.values()[0]
                collecttargetitmes_objects = PlatformTargetItem.objects.filter(platform_id=collecttarget_objects_value['platform_id'])
                collecttargetitmes_values = collecttargetitmes_objects.values()
                for collecttargetitmes_value in collecttargetitmes_values:
                    collecttargetitems_datas.append(collecttargetitmes_value)
                return JsonResponse(data={'success': True, 'data': collecttargetitems_datas,'platform_id':collecttarget_objects_value['platform_id']})
            else:
                return JsonResponse(data={'success': True, 'data': []})
        except:
            return JsonResponse(status=400, data={'success': False})

    # @login_required
    def put(self, request):
        """
        PlatformTargetItem update api
        """
        try:
            collecttargetitem_list = JSONParser().parse(request)
            for i,collecttargetitem_object in enumerate(collecttargetitem_list):

                collecttargetitem_data = PlatformTargetItem.objects.filter(platform_id=collecttargetitem_object["platform"])[i]
                collecttargetitem_serializer = PlatformTargetItemSerializer(collecttargetitem_data, data=collecttargetitem_object)
                if collecttargetitem_serializer.is_valid():
                    collecttargetitem_serializer.save()
                else:
                    return JsonResponse(data={'success': False,'data': collecttargetitem_serializer.errors}, status=400)
            return JsonResponse(data={'success': True}, status=status.HTTP_201_CREATED)
        except:
            return JsonResponse(data={'success': False}, status=400)


class DataReportAPI(APIView):
    def get(self, request):
        """
        Data-Report read api
        """
        platform = request.GET.get('platform', None)
        type = request.GET.get('type', None)
        start_date = request.GET.get('start_date', None)
        end_date = request.GET.get('end_date', None)

        #artist name
        artist_objects = Artist.objects.filter(active=1)
        artist_objects_values = artist_objects.values()
        artist_list = []
        for a in artist_objects_values:
            artist_list.append(a['name'])

        #platform target names
        platform_id = Platform.objects.get(name = platform).id
        collecttargets = CollectTarget.objects.filter(platform = platform_id)
        collecttargets = collecttargets.values()
        platform_list = set()
        for collecttarget in collecttargets:
            platform_objects = CollectTargetItem.objects.filter(collect_target_id = collecttarget['id'])
            platform_objects_values = platform_objects.values()
            for p in platform_objects_values:
                if p["target_name"] in platform_list:
                    continue
                platform_list.add(p["target_name"])
        platform_list = list(platform_list)

        #플랫폼 헤더 정보 순서와 db 칼럼 저장 순서 싱크 맞추기
        platform_header = []
        objects = DataModels[platform].objects.all()
        objects_values = objects.values()
        obj_datas = []
        for v in objects_values:
            obj_datas.append(v)
        if len(obj_datas) > 0:
            key_list = list(obj_datas[0].keys())
        else:
            key_list = []
        for key in key_list:
            if key in platform_list:
                platform_header.append(key)
            else:
                continue


        try:
            if type == "누적":
                start_date_dateobject = datetime.datetime.strptime(start_date, '%Y-%m-%d')
                filter_objects = DataModels[platform].objects.filter(reserved_date__year=start_date_dateobject.year,
                    reserved_date__month=start_date_dateobject.month, reserved_date__day=start_date_dateobject.day)
                if filter_objects.exists():
                    filter_objects_values=filter_objects.values()
                    filter_datas=[]
                    for filter_value in filter_objects_values:
                        filter_datas.append(filter_value)
                    return JsonResponse(data={'success': True, 'data': filter_datas,'artists':artist_list,'platform':platform_header})
                else:
                    crawling_artist_list = [] 
                    objects = DataModels[platform].objects.all()
                    objects_value = objects.values()
                    for val in objects_value:
                        crawling_artist_list.append(val['artist'])
                    #datename = '%s-%s-%s'%(start_date_dateobject.year, start_date_dateobject.month, start_date_dateobject.day)
                    return JsonResponse(status=200, data={'success': True, 'data':'no data','artists':artist_list,'platform':platform_header,'crawling_artist_list':crawling_artist_list})
            elif type == "기간별":
                # 전날 값을 구함
                start_date_dateobject=datetime.datetime.strptime(start_date, '%Y-%m-%d').date() - datetime.timedelta(1)
                end_date_dateobject=datetime.datetime.strptime(end_date, '%Y-%m-%d').date()
                filter_objects_start=DataModels[platform].objects.filter(reserved_date__year=start_date_dateobject.year,
                    reserved_date__month=start_date_dateobject.month, reserved_date__day=start_date_dateobject.day)
                filter_objects_end=DataModels[platform].objects.filter(reserved_date__year=end_date_dateobject.year,
                    reserved_date__month=end_date_dateobject.month, reserved_date__day=end_date_dateobject.day)
                filter_datas_total=[]
                # 둘 다 존재할 때
                if filter_objects_start.exists() and filter_objects_end.exists():
                    filter_objects_start_values=filter_objects_start.values()
                    model_fields = DataModels[platform]._meta.fields
                    model_fields_name = []
                    artist_datas = set()
                    for model_field in model_fields:
                        model_fields_name.append(model_field.name)
                    values_len = len(filter_objects_start_values)
                    for i in range(values_len):
                        # 이미 넣은 데이터면 pass
                        if filter_objects_start_values[i]["artist"] in artist_datas:
                            continue
                        artist_datas.add(filter_objects_start_values[i]["artist"])
                        # id랑 artist, date 빼고 보내주기
                        data_json = {}
                        # 현재 보고 있는 거랑 맞는 끝 날짜를 가져오기
                        filter_artist_end=DataModels[platform].objects.filter(reserved_date__year=end_date_dateobject.year,
                            reserved_date__month=end_date_dateobject.month, reserved_date__day=end_date_dateobject.day,
                            artist = filter_objects_start_values[i]["artist"])
                        filter_artist_end = filter_artist_end.values()
                        if not filter_artist_end.exists():
                            continue
                        filter_artist_end = filter_artist_end[0]
                        for field_name in model_fields_name:
                            if field_name != "id" and field_name != "artist" and field_name != "user_created" and field_name != "recorded_date" and field_name != "platform" and field_name != "url" :
                                if filter_artist_end[field_name] is not None and filter_objects_start_values[i][field_name] is not None:
                                    data_json[field_name] = filter_artist_end[field_name] - filter_objects_start_values[i][field_name]
                                elif filter_artist_end[field_name] is not None: #앞의 날짜를 0으로 처리한 형태
                                    data_json[field_name] = filter_artist_end[field_name]
                                else:
                                    data_json[field_name] = 0
                            else: #숫자 아닌 다른 정보들
                                data_json[field_name] = filter_objects_start_values[i][field_name]
                        filter_datas_total.append(data_json)
                    return JsonResponse(data={'success': True, 'data': filter_datas_total,'artists':artist_list,'platform':platform_header})
                elif not filter_objects_start.exists() and filter_objects_end.exists():
                    # 시작날짜의 데이터가 존재하지 않고 끝날짜의 데이터만 존재할 때
                    # 0으로 해서 계산
                    filter_objects_end_values=filter_objects_end.values()
                    model_fields = DataModels[platform]._meta.fields
                    model_fields_name = []
                    artist_datas = set()
                    for model_field in model_fields:
                        model_fields_name.append(model_field.name)
                    values_len = len(filter_objects_end_values)
                    for i in range(values_len):
                        # 이미 넣은 데이터면 pass
                        if filter_objects_end_values[i]["artist"] in artist_datas:
                            continue
                        artist_datas.add(filter_objects_end_values[i]["artist"])
                        data_json = {}
                        for field_name in model_fields_name:
                            if field_name != "id" and field_name != "artist" and field_name != "user_created" and field_name != "recorded_date" and field_name != "platform" and field_name != "url" :
                                # None때문에 오류나면 비워놓고 보내기
                                if filter_objects_end_values[i][field_name] is not None:
                                    data_json[field_name] = filter_objects_end_values[i][field_name]
                            else:
                                data_json[field_name] = filter_objects_end_values[i][field_name]
                        filter_datas_total.append(data_json)
                    return JsonResponse(data={'success': True, 'data': filter_datas_total,'artists':artist_list,'platform':platform_header})
                # 끝날짜의 데이터가 아예 존재하지 않을 때
                else:
                    datename = '%s-%s-%s'%(end_date_dateobject.year, end_date_dateobject.month, end_date_dateobject.day)
                    return JsonResponse(status=400, data={'success': False, 'data': datename})
            else:
                if DataModels[platform].objects.exists():
                    platform_queryset_values = DataModels[platform].objects.values()
                    platform_datas = []
                    for queryset_value in platform_queryset_values:
                        platform_datas.append(queryset_value)
                    return JsonResponse(data={'success': True, 'data': platform_datas,'artists':artist_list,'platform':platform_header})
                else:
                    return JsonResponse(status=400, data={'success': False, 'data': 'there is no data'})
        except:
            return JsonResponse(status=400, data={'success': False})
    
    def post(self, request):
        """
        Data-Report update api
        """
        update_data_object = JSONParser().parse(request)
        start_date = update_data_object[len(update_data_object)-1]['start_date']
        platform = update_data_object[len(update_data_object)-1]['platform_name']

        # artist name
        artist_objects = Artist.objects.filter(active=1)

        artist_objects_values = artist_objects.values()
        artist_list = []
        for a in artist_objects_values:
            artist_list.append(a['name'])

        #crawled artist list
        a_objects = DataModels[platform].objects.all()
        a_objects_values = a_objects.values()
        a_list = []
        for val in a_objects_values:
            a_list.append(val['artist'])
        

        #platform target names
        platform_id = Platform.objects.get(name = platform).id
        collecttargets = CollectTarget.objects.filter(platform = platform_id)
        collecttargets = collecttargets.values()
        platform_list = set()
        for collecttarget in collecttargets:
            platform_objects = CollectTargetItem.objects.filter(collect_target_id = collecttarget['id'])
            platform_objects_values = platform_objects.values()
            for p in platform_objects_values:
                if p["target_name"] in platform_list:
                    continue
                platform_list.add(p["target_name"])
        platform_list = list(platform_list)

        #플랫폼 헤더 정보 순서와 db 칼럼 저장 순서 싱크 맞추기
        platform_header = []
        objects = DataModels[platform].objects.all()
        objects_values = objects.values()
        obj_datas = []
        for v in objects_values:
            obj_datas.append(v)
        key_list = list(obj_datas[0].keys())

        for key in key_list:
            if key in platform_list:
                platform_header.append(key)
            else:
                continue

        try:
            start_date_dateobject = datetime.datetime.strptime(start_date, "%Y-%m-%d")
            for index, element in enumerate(update_data_object):
                if index == len(update_data_object)-1:
                    break
                obj = DataModels[platform].objects.filter(artist=element['artist'], reserved_date = start_date)
                print(element['artist'])
                obj = DataModels[platform].objects.get(artist=artist,reserved_date__year=start_date_dateobject.year,
                reserved_date__month=start_date_dateobject.month, reserved_date__day=start_date_dateobject.day)

                past_uploads = obj["uploads"]
                past_subscribers = obj["subscribers"]
                past_views = obj["views"]
                past_videos = obj["videos"]
                past_likes = obj["likes"]
                past_plays = obj["plays"]
                past_followers = obj["followers"]
                past_twits = obj["twits"]
                past_weverses = obj["weverses"]
                past_listens = obj["monthly_listens"]
                past_listeners = obj["listener"]
                past_streams = obj["streams"]

                if obj: #처음부터 크롤링 잘 된 경우
                    if platform == 'youtube':
                        if element['target'] == 'uploads':
                            if past_uploads != element['current']:
                                userlogger.info(f"[CHANGE]: {artist}'s uploads of {platform} [{past_uploads}-> {element['current']}]")
                            obj.update(uploads=element['current'])
                        elif element['target'] == 'subscribers':
                            if past_subscribers != element['current']:
                                userlogger.info(f"[CHANGE]: {artist}'s subscribers of {platform} [{past_subscribers}-> {element['current']}]")
                            obj.update(subscribers=element['current'])
                        elif element['target'] == 'views':
                            if past_views != element['current']:
                                userlogger.info(f"[CHANGE]: {artist}'s views of {platform} [{past_views}-> {element['current']}]")
                            obj.update(views=element['current'])
                        else:
                            obj.update(user_created=element['comma_current'])

                    elif platform == 'vlive':
                        if element['target'] == 'members':
                            obj.update(members=element['current'])
                        elif element['target'] == 'videos':
                            if past_videos != element['current']:
                                userlogger.info(f"[CHANGE]: {artist}'s videos of {platform} [{past_videos}-> {element['current']}]")
                            obj.update(videos=element['current'])
                        elif element['target'] == 'likes':
                            if past_likes != element['current']:
                                userlogger.info(f"[CHANGE]: {artist}'s likes of {platform} [{past_likes}-> {element['current']}]")
                            obj.update(likes=element['current'])
                        else:
                            if past_plays != element['current']:
                                userlogger.info(f"[CHANGE]: {artist}'s plays of {platform} [{past_plays}-> {element['current']}]")
                            obj.update(plays=element['current'])
    
                    elif platform == 'instagram' or platform=='facebook':
                        if past_followers != element['current']:
                            userlogger.info(f"[CHANGE]: {artist}'s uploads of {platform} [{past_followers}-> {element['current']}]")
                        obj.update(followers = element['current'])

                    elif platform == 'twitter' or platform=='twitter2':
                        if element['target'] == 'followers':
                            if past_followers != element['current']:
                                userlogger.info(f"[CHANGE]: {artist}'s followers of {platform} [{past_followers}-> {element['current']}]")
                            obj.update(followers=element['current'])
                        elif element['target'] == 'twits':
                            if past_twits != element['current']:
                                userlogger.info(f"[CHANGE]: {artist}'s twits of {platform} [{past_twits}-> {element['current']}]")
                            obj.update(twits=element['current'])
                        else:
                            obj.update(user_created=element['comma_current'])

                    elif platform == 'tiktok':
                        if element['target'] == 'followers':
                            if past_followers != element['current']:
                                userlogger.info(f"[CHANGE]: {artist}'s followers of {platform} [{past_followers}-> {element['current']}]")
                            obj.update(followers=element['current'])
                        elif element['target'] == 'uploads':
                            if past_uploads != element['current']:
                                userlogger.info(f"[CHANGE]: {artist}'s uploads of {platform} [{past_uploads}-> {element['current']}]")
                            obj.update(uploads=element['current'])
                        else:
                            if past_likes != element['current']:
                                userlogger.info(f"[CHANGE]: {artist}'s likes of {platform} [{past_likes}-> {element['current']}]")
                            obj.update(likes=element['current'])

                    elif platform == 'weverse':
                        if past_weverses != element['current']:
                            userlogger.info(f"[CHANGE]: {artist}'s uploads of {platform} [{past_weverses}-> {element['current']}]")
                        obj.update(weverses= element['current'])

                    elif platform == 'spotify':
                        if element['target'] == 'monthly_listens':
                            if past_listens != element['current']:
                                userlogger.info(f"[CHANGE]: {artist}'s listens of {platform} [{past_listens}-> {element['current']}]")
                            obj.update(monthly_listens=element['current'])
                        else:
                            if past_followers != element['current']:
                                userlogger.info(f"[CHANGE]: {artist}'s followers of {platform} [{past_followers}-> {element['current']}]")
                            obj.update(followers=element['current'])
                    
                    elif platform == 'melon':
                        if element['target'] == 'listeners':
                            if past_listeners != element['current']:
                                userlogger.info(f"[CHANGE]: {artist}'s listeners of {platform} [{past_listeners}-> {element['current']}]")
                            obj.update(listeners=element['current'])
                        else:
                            if past_streams != element['current']:
                                userlogger.info(f"[CHANGE]: {artist}'s streams of {platform} [{past_streams}-> {element['current']}]")
                            obj.update(streams=element['current'])
                else:
                    if element['artist'] in a_list:
                        if platform == 'youtube':
                            if element['target'] == 'uploads':
                                instance = DataModels[platform](uploads=element['current'],reserved_date = start_date)
                                instance.save()
                            elif element['target'] == 'subscribers':
                                instance = DataModels[platform](subscribers=element['current'],reserved_date=start_date)
                                instance.save()
                            elif element['target'] == 'views':
                                instance = DataModels[platform](views=element['current'],reserved_date = start_date)
                                instance.save()
                            else:
                                instance = DataModels[platform](user_created=element['comma_current'])
                                instance.save()
                        elif platform == 'vlive':
                            if element['target'] == 'members':
                                instance = DataModels[platform](members=element['current'],reserved_date = start_date)
                                instance.save()
                            elif element['target'] == 'videos':
                                instance = DataModels[platform](videos=element['current'],reserved_date = start_date)
                                instance.save()
                            elif element['target'] == 'likes':
                                instance = DataModels[platform](likes=element['current'],reserved_date = start_date)
                                instance.save()
                            else:
                                instance = DataModels[platform](plays=element['current'],reserved_date = start_date)
                                instance.save()
    
                        elif platform == 'instagram' or platform=='facebook':
                            instance = DataModels[platform](followers = element['current'],reserved_date = start_date)
                            instance.save()
                        elif platform == 'twitter' or platform=='twitter2':
                            if element['target'] == 'followers':
                                instance = DataModels[platform](followers=element['current'],reserved_date = start_date)
                                instance.save()
                            elif element['target'] == 'twits':
                                instance = DataModels[platform](twits=element['current'],reserved_date = start_date)
                                instance.save()
                            else:
                                instance = DataModels[platform](user_created=element['comma_current'],reserved_date = start_date)
                                instance.save()
                        elif platform == 'tiktok':
                            if element['target'] == 'followers':
                                instance = DataModels[platform](followers=element['current'],reserved_date = start_date)
                                instance.save()
                            elif element['target'] == 'uploads':
                                instance = DataModels[platform](uploads=element['current'],reserved_date = start_date)
                                instance.save()
                            else:
                                instance = DataModels[platform](likes=element['current'],reserved_date = start_date)
                                instance.save()
                        elif platform == 'weverse':
                            instance = DataModels[platform](weverses= element['current'],reserved_date = start_date)
                            instance.save()
                        elif platform == 'spotify':
                            if element['target'] == 'monthly_listens':
                                instance = DataModels[platform](monthly_listens=element['current'],reserved_date = start_date)
                                instance.save()
                            else:
                                instance = DataModels[platform](followers=element['current'],reserved_date = start_date)
                                instance.save()
                        elif platform == 'melon':
                            if element['target'] == 'listeners':
                                instance = DataModels[platform](listeners=element['current'],reserved_date = start_date)
                                instance.save()
                            else:
                                instance = DataModels[platform](streams=element['current'],reserved_date = start_date)
                                instance.save()
                    else:
                        pass

            filter_objects = DataModels[platform].objects.filter(reserved_date__year=start_date_dateobject.year,
                reserved_date__month=start_date_dateobject.month, reserved_date__day=start_date_dateobject.day)
            if filter_objects.exists():
                filter_objects_values=filter_objects.values()
                filter_datas=[]

                crawling_artist_list = [] 
                objects = DataModels[platform].objects.all()
                objects_value = objects.values()
                for val in objects_value:
                    crawling_artist_list.append(val['artist'])
                   
                for filter_value in filter_objects_values:
                    filter_datas.append(filter_value)
                return JsonResponse(data={'success': True, 'data': filter_datas,'artists':artist_list,'platform':platform_header,'crawling_artist_list':crawling_artist_list})
            else:
                crawling_artist_list = [] 
                objects = DataModels[platform].objects.all()
                objects_value = objects.values()
                for val in objects_value:
                    crawling_artist_list.append(val['artist'])
                #datename = '%s-%s-%s'%(start_date_dateobject.year, start_date_dateobject.month, start_date_dateobject.day)
                return JsonResponse(status=200, data={'success': True, 'data':'no data','artists':artist_list,'platform':platform_header,'crawling_artist_list':crawling_artist_list})
        except:
            return JsonResponse(status=400, data={'success': False})
